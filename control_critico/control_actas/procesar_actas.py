from .excel_utils import obtener_columnas
from .config import ACTIVIDADES_CRITICAS

import os
import math
import unicodedata
import re
from openpyxl import load_workbook
from openpyxl.styles import Font


def normalizar(texto):
    if texto is None:
        return ""
    s = str(texto).upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalizar_unidad(u):
    if u is None:
        return ""
    s = str(u).strip().upper()
    s = s.replace("M³", "M3").replace("M^3", "M3")
    s = s.replace("M²", "M2").replace("M^2", "M2")
    s = re.sub(r"\s+", "", s)
    mapa = {"UND": "UN", "UNID": "UN", "UNIDAD": "UN", "U": "UN", "ML": "M", "M.L": "M"}
    return mapa.get(s, s)


# Diccionario normalizado para búsqueda rápida
CRITICOS_NORM = {}
for k, v in ACTIVIDADES_CRITICAS.items():
    kn = normalizar(k)
    if not kn:
        continue
    try:
        CRITICOS_NORM[kn] = float(v)
    except Exception:
        continue


def _buscar_critico(desc_norm: str):
    """
    Retorna (crit_key_norm, precio_ref) usando 'in'.
    Si hay varios matches, escoge el más largo para evitar matches muy genéricos.
    """
    best_k = None
    best_len = -1
    best_precio = None
    for k_norm, precio in CRITICOS_NORM.items():
        if k_norm in desc_norm:
            if len(k_norm) > best_len:
                best_len = len(k_norm)
                best_k = k_norm
                best_precio = precio
    return best_k, best_precio


def _clasificar_familia(desc_norm: str) -> str | None:
    """
    Clasificación directa por palabra clave en la DESCRIPCIÓN normalizada.
    Reglas:
      - estamp -> CONCRETO_ESTAMPADO
      - mr     -> CONCRETO_MR
      - excav  -> EXCAVACIONES
      - rellen -> RELLENOS
    """
    d = f" {desc_norm} "

    if " ESTAMP" in d:
        return "CONCRETO_ESTAMPADO"

    # MR como palabra
    if re.search(r"\bMR\b", d):
        return "CONCRETO_MR"

    if " EXCAV" in d:
        return "EXCAVACIONES"

    if " RELLEN" in d:
        return "RELLENOS"

    return None


def _extraer_cantidades_por_familia(ws_vals, columnas: dict) -> dict[str, list[dict]]:
    """
    Lee CORTE y arma tablas (solo cantidades) por familia.
    Devuelve dict: {FAMILIA: [ {item, descripcion, un, cantidad}, ... ] }
    """
    col_item = columnas.get("ÍTEM", "A")
    col_desc = columnas.get("DESCRIPCIÓN", "B")
    col_un = columnas.get("UN", "C")      
    col_cantidad = "G"                   

    out: dict[str, list[dict]] = {
        "RELLENOS": [],
        "EXCAVACIONES": [],
        "CONCRETO_MR": [],
        "CONCRETO_ESTAMPADO": [],
    }

    fila_inicio = 10
    for fila in range(fila_inicio, ws_vals.max_row + 1):
        item = str(ws_vals[f"{col_item}{fila}"].value or "").strip()
        desc_raw = ws_vals[f"{col_desc}{fila}"].value
        un_raw = ws_vals[f"{col_un}{fila}"].value

        if not item or not desc_raw:
            continue

        descripcion = str(desc_raw).strip()
        unidad = normalizar_unidad(un_raw)

        # cantidad
        cantidad_cell = ws_vals[f"{col_cantidad}{fila}"].value
        try:
            cantidad = float(cantidad_cell)
        except (TypeError, ValueError):
            continue

        if cantidad == 0 or (isinstance(cantidad, float) and math.isnan(cantidad)):
            continue

        desc_norm = normalizar(descripcion)
        if not desc_norm:
            continue

        familia = _clasificar_familia(desc_norm)
        if not familia:
            continue

        out[familia].append({
            "item": item,
            "descripcion": descripcion,
            "un": unidad,
            "cantidad": cantidad,
        })

    return out


def _crear_hoja_cuadro_cantidades(wb, tablas: dict[str, list[dict]], nombre="CUADRO_CANTIDADES"):
    """
    Crea/reemplaza una sola hoja con 4 columnas:
    Excavacione | Rellenos | Concreto MR | Concreto estampado
    y debajo solo las cantidades.
    """
    if nombre in wb.sheetnames:
        wb.remove(wb[nombre])

    ws = wb.create_sheet(title=nombre)

    ws["B1"] = "Excavacione"
    ws["C1"] = "Rellenos"
    ws["D1"] = "Concreto MR"
    ws["E1"] = "Concreto estampado"

    col_map = {
        "EXCAVACIONES": "B",
        "RELLENOS": "C",
        "CONCRETO_MR": "D",
        "CONCRETO_ESTAMPADO": "E",
    }

    for familia, col in col_map.items():
        cantidades = [r["cantidad"] for r in tablas.get(familia, [])]
        for i, val in enumerate(cantidades, start=2):
            ws[f"{col}{i}"] = val

    max_len = 0
    for familia in col_map.keys():
        max_len = max(max_len, len(tablas.get(familia, [])))

    fila_total = 2 + max_len
    ws[f"A{fila_total}"] = "TOTAL"

    for col in ["B", "C", "D", "E"]:
        if max_len > 0:
            ws[f"{col}{fila_total}"] = f"=SUM({col}2:{col}{fila_total-1})"
        else:
            ws[f"{col}{fila_total}"] = 0


def revisar_acta(
    path_archivo: str,
    anio_actual: int,
    mes_nombre: str,
    carpeta_salida_mes: str,
    base_registro: list,
    base_cantidades: list | None = None,  
):
    try:
        wb = load_workbook(path_archivo, data_only=False)
        wb_vals = load_workbook(path_archivo, data_only=True)

        hoja_corte = None
        for n in ["CORTE", "Corte", "corte", "Corte "]:
            if n in wb.sheetnames:
                hoja_corte = n
                break

        if not hoja_corte:
            return

        ws = wb[hoja_corte]
        ws_vals = wb_vals[hoja_corte]

    except Exception:
        return

    nombre_contratista = ws["B6"].value or ws["C6"].value or ws["D6"].value or "SIN NOMBRE"
    columnas = obtener_columnas(ws)

    # Contadores de cantidades
    totales_cant = {
        "Excavaciones": 0.0,
        "Rellenos": 0.0,
        "Concreto MR": 0.0,
        "Concreto estampado": 0.0,
    }

    col_item = columnas.get("ÍTEM", "A")
    col_desc = columnas.get("DESCRIPCIÓN", "B")
    col_un = columnas.get("UN", "C")       

    col_valor = columnas.get("VALOR UNITARIO")
    if not col_valor:
        for key, letra in columnas.items():
            if "VALOR UNITARIO" in key:
                col_valor = letra
                break
    if not col_valor:
        return

    col_cantidad = "G"  

    for fila in range(10, ws.max_row + 1):
        item = str(ws[f"{col_item}{fila}"].value or "").strip()
        desc_raw = ws[f"{col_desc}{fila}"].value
        un_raw = ws[f"{col_un}{fila}"].value

        if not item or not desc_raw:
            continue

        descripcion = str(desc_raw).strip()
        desc_norm = normalizar(descripcion)

        # Filtros
        if "MANO DE OBRA" in desc_norm or "PEA" in desc_norm:
            continue

        # ==========================================
        # ✅ CANTIDADES POR CATEGORÍA (NO DEPENDEN DE CRÍTICO)
        # ==========================================
        try:
            cantidad = float(ws_vals[f"{col_cantidad}{fila}"].value)
        except Exception:
            cantidad = 0.0

        if cantidad and not (isinstance(cantidad, float) and math.isnan(cantidad)):
            if "EXCAV" in desc_norm:
                totales_cant["Excavaciones"] += float(cantidad)
            if "RELLEN" in desc_norm:
                totales_cant["Rellenos"] += float(cantidad)
            if re.search(r"\bMR\b", desc_norm):
                totales_cant["Concreto MR"] += float(cantidad)
            if "ESTAMP" in desc_norm:
                totales_cant["Concreto estampado"] += float(cantidad)

        # ==========================================
        #  MODO CRÍTICO (precios): SOLO si matchea actividad crítica
        # ==========================================
        _, precio_ref = _buscar_critico(desc_norm)
        if precio_ref is None:
            continue

        # Leer valor unitario
        try:
            valor = float(
                str(ws_vals[f"{col_valor}{fila}"].value)
                .replace("$", "")
                .replace(",", "")
            )
        except Exception:
            continue

        # Si no hay cantidad válida, no registramos error (pero ya la contamos arriba)
        if (not cantidad) or (isinstance(cantidad, float) and math.isnan(cantidad)):
            continue

        celda_valor = ws[f"{col_valor}{fila}"]
        diff = valor - precio_ref

        if abs(diff) > 1:
            if diff > 0:
                celda_valor.font = Font(color="FFFF0000")  # rojo
            else:
                celda_valor.font = Font(color="FF0000FF")  # azul

            base_registro.append({
                "anio": anio_actual,
                "mes": mes_nombre,
                "archivo": os.path.basename(path_archivo),
                "contratista": nombre_contratista,
                "item": item,
                "descripcion": descripcion,
                "valor_unitario_original": valor,
                "un": normalizar_unidad(un_raw),
                "valor_pactado": precio_ref,
                "cantidad_presenta": cantidad,
                "valor_ajustado": precio_ref * cantidad,
                "modo": "CRITICO",
            })

    salida = os.path.join(
        carpeta_salida_mes,
        os.path.basename(path_archivo).replace(".xlsx", "_verificado.xlsx")
    )

    # ✅ Guardar totales por acta (igual lógica del normal)
    if base_cantidades is not None:
        base_cantidades.append({
            "anio": anio_actual,
            "mes": mes_nombre,
            "archivo": os.path.basename(path_archivo),
            "contratista": nombre_contratista,
            "Excavaciones": totales_cant["Excavaciones"],
            "Rellenos": totales_cant["Rellenos"],
            "Concreto MR": totales_cant["Concreto MR"],
            "Concreto estampado": totales_cant["Concreto estampado"],
            "modo": "CRITICO",
        })

    # ✅ Hoja de cuadro de cantidades (misma lógica del normal)
    tablas = _extraer_cantidades_por_familia(ws_vals, columnas)
    _crear_hoja_cuadro_cantidades(wb, tablas, nombre="CUADRO_CANTIDADES")

    wb.save(salida)












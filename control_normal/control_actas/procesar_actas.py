from .excel_utils import obtener_columnas

import os
import math
import unicodedata
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from .config import ACTIVIDADES_CRITICAS



def normalizar(texto):
    if texto is None:
        return None
    s = str(texto).lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalizar_unidad(u):
    if u is None:
        return ""
    s = str(u).strip().upper()
    s = s.replace("M³", "M3").replace("M^3", "M3")
    s = s.replace("M²", "M2").replace("M^2", "M2")
    s = re.sub(r"\s+", "", s)
    mapa = {"UND": "UN", "UNID": "UN", "UNIDAD": "UN", "U": "UN", "ML": "M", "M.L": "M"}
    return mapa.get(s, s)


# --- MODO CRÍTICO: keyword -> precio_ref (solo diccionario) ---
CRIT_MAP = {}
for k, v in ACTIVIDADES_CRITICAS.items():
    kn = normalizar(k)
    if not kn:
        continue
    try:
        CRIT_MAP[kn] = float(v)
    except Exception:
        continue


def _buscar_critico(desc_norm: str):
    """
    Retorna (crit_key_norm, precio_ref) usando 'in'.
    Si hay varios matches, escoge el más largo para evitar matches muy genéricos.
    """
    best_k = None
    best_len = -1
    best_precio = None
    for k_norm, precio in CRIT_MAP.items():
        if k_norm in desc_norm:
            if len(k_norm) > best_len:
                best_len = len(k_norm)
                best_k = k_norm
                best_precio = precio
    return best_k, best_precio

def _clasificar_familia(desc_norm: str) -> str | None:
    """
    Clasificación directa por palabra clave en la DESCRIPCIÓN normalizada.
    Reglas:
      - estamp -> CONCRETO_ESTAMPADO
      - mr     -> CONCRETO_MR
      - excav  -> EXCAVACIONES
      - rellen -> RELLENOS
    """
    d = f" {desc_norm} "  # para buscar palabras con borde

    if " estamp" in d:
        return "CONCRETO_ESTAMPADO"

    # MR como palabra (evita capturar cosas raras tipo "mra" accidentalmente)
    if re.search(r"\bmr\b", d):
        return "CONCRETO_MR"

    if " excav" in d:
        return "EXCAVACIONES"

    if " rellen" in d:
        return "RELLENOS"

    return None


def _extraer_cantidades_por_familia(ws_vals, columnas: dict) -> dict[str, list[dict]]:
    """
    Lee CORTE y arma tablas (solo cantidades) por familia.
    Devuelve dict: {FAMILIA: [ {item, descripcion, un, cantidad}, ... ] }
    """
    col_item = columnas.get("ÍTEM", "A")
    col_desc = columnas.get("DESCRIPCIÓN", "B")
    col_un = columnas.get("UN", "D")
    col_cantidad = "I"  # fija según tu archivo

    out: dict[str, list[dict]] = {
        "RELLENOS": [],
        "EXCAVACIONES": [],
        "CONCRETO_MR": [],
        "CONCRETO_ESTAMPADO": [],
    }

    fila_inicio = 10
    for fila in range(fila_inicio, ws_vals.max_row + 1):
        item = str(ws_vals[f"{col_item}{fila}"].value or "").strip()
        desc_raw = ws_vals[f"{col_desc}{fila}"].value
        un_raw = ws_vals[f"{col_un}{fila}"].value

        if not item or not desc_raw:
            continue

        descripcion = str(desc_raw).strip()
        unidad = normalizar_unidad(un_raw)

        # cantidad
        cantidad_cell = ws_vals[f"{col_cantidad}{fila}"].value
        try:
            cantidad = float(cantidad_cell)
        except (TypeError, ValueError):
            continue

        if cantidad == 0 or (isinstance(cantidad, float) and math.isnan(cantidad)):
            continue

        # normaliza y clasifica
        desc_norm = normalizar(descripcion)
        if not desc_norm:
            continue

        familia = _clasificar_familia(desc_norm)
        if not familia:
            continue

        out[familia].append({
            "item": item,
            "descripcion": descripcion,
            "un": unidad,
            "cantidad": cantidad,
        })

    return out

def _crear_hoja_cuadro_cantidades(wb, tablas: dict[str, list[dict]], nombre="CUADRO_CANTIDADES"):
    """
    Crea/reemplaza una sola hoja con 4 columnas:
    Excavacione | Rellenos | Concreto MR | Concreto estampado
    y debajo solo las cantidades.
    """

    # 1) borrar si existe
    if nombre in wb.sheetnames:
        wb.remove(wb[nombre])

    ws = wb.create_sheet(title=nombre)

    # 2) encabezados en B..E (como tu imagen)
    ws["B1"] = "Excavacione"
    ws["C1"] = "Rellenos"
    ws["D1"] = "Concreto MR"
    ws["E1"] = "Concreto estampado"

    # 3) listas de cantidades por familia
    col_map = {
        "EXCAVACIONES": "B",
        "RELLENOS": "C",
        "CONCRETO_MR": "D",
        "CONCRETO_ESTAMPADO": "E",
    }

    # 4) pegar cantidades desde fila 2 hacia abajo
    for familia, col in col_map.items():
        cantidades = [r["cantidad"] for r in tablas.get(familia, [])]

        for i, val in enumerate(cantidades, start=2):
            ws[f"{col}{i}"] = val
    # 5) Fila TOTAL: sumar cada columna (B..E)
    max_len = 0
    for familia in col_map.keys():
        max_len = max(max_len, len(tablas.get(familia, [])))

    fila_total = 2 + max_len  # debajo del último dato

    # (opcional) etiqueta en A
    ws[f"A{fila_total}"] = "TOTAL"

    for col in ["B", "C", "D", "E"]:
        if max_len > 0:
            ws[f"{col}{fila_total}"] = f"=SUM({col}2:{col}{fila_total-1})"
        else:
            ws[f"{col}{fila_total}"] = 0


def revisar_acta(
    path_archivo: str,
    valores_referencia: dict,   # dict[(desc_norm, un_norm)] -> precio_ref  (solo se usa en modo normal)
    anio_actual: int,
    mes_nombre: str,
    carpeta_salida_mes: str,
    base_registro: list,
    base_cantidades: list,      # ✅ NUEVO: para guardar totales por contratista/categoría
    modo_critico: bool = False,
):
    # ✅ NUEVO: acumuladores por acta
    totales_cant = {
        "Excavaciones": 0.0,
        "Rellenos": 0.0,
        "Concreto MR": 0.0,
        "Concreto estampado": 0.0,
    }
        # 🐞 DEBUG CONTADORES
    debug = {
        "filas_iteradas": 0,
        "sin_item_o_desc": 0,
        "valor_unitario_none": 0,
        "valor_unitario_no_num": 0,
        "cantidad_no_num": 0,
        "sin_valor_referencia": 0,
        "registros_guardados": 0,
    }


    try:
        wb = load_workbook(path_archivo, data_only=False)     # para escribir estilos
        wb_vals = load_workbook(path_archivo, data_only=True) # para leer valores calculados

        hoja_corte = None
        for nombre in ["CORTE", "Corte", "Corte ", "corte"]:
            if nombre in wb.sheetnames:
                hoja_corte = nombre
                break

        if not hoja_corte:
            print(f"❌ No se encontró la hoja 'CORTE' o 'Corte' en {path_archivo}")
            return

        ws = wb[hoja_corte]
        ws_vals = wb_vals[hoja_corte]
    except Exception as e:
        print(f"❌ Error al abrir {path_archivo}: {e}")
        return

    nombre_contratista = ws["C6"].value or ws["D6"].value or "SIN NOMBRE"
    columnas = obtener_columnas(ws)

    col_item = columnas.get("ÍTEM", "A")
    col_desc = columnas.get("DESCRIPCIÓN", "B")
    col_un = columnas.get("UN", "D")  # en tu archivo suele ser D

    col_valor = columnas.get("VALOR UNITARIO", None)
    if not col_valor:
        for key, letra in columnas.items():
            if "VALOR UNITARIO" in key:
                col_valor = letra
                break
    if not col_valor:
        print(f"⚠ No se encontró columna 'VALOR UNITARIO' en {path_archivo}")
        return

    col_cantidad_presenta = "I"  # fija

    fila_inicio = 10
    for fila in range(fila_inicio, ws.max_row + 1):
        debug["filas_iteradas"] += 1
        item = str(ws[f"{col_item}{fila}"].value or "").strip()
        desc_raw = ws[f"{col_desc}{fila}"].value
        un_raw = ws[f"{col_un}{fila}"].value

        if not item or not desc_raw:
            debug["sin_item_o_desc"] += 1
            continue


        descripcion = str(desc_raw).strip()
        unidad = normalizar_unidad(un_raw)

        # filtros
        desc_upper = descripcion.upper()
        if (
            "MANO DE OBRA" in desc_upper
            or "MR45" in item.upper()
            or "PEA" in desc_upper
        ):
            continue

        desc_norm = normalizar(descripcion)
        if not desc_norm:
            continue

        # leer valor unitario desde wb_vals (por fórmulas)
        valor_cell = ws_vals[f"{col_valor}{fila}"].value
        
        if valor_cell is None:
            debug["valor_unitario_none"] += 1
            continue
        
        try:
            valor = float(str(valor_cell).replace("$", "").replace(",", ""))
        except Exception:
            debug["valor_unitario_no_num"] += 1
            continue

        # cantidad presenta (columna I)
        cantidad_presenta = ws_vals[f"{col_cantidad_presenta}{fila}"].value
        try:
            cantidad_num = float(cantidad_presenta)
        except (TypeError, ValueError):
            debug["cantidad_no_num"] += 1
            cantidad_num = 0.0


        es_nan = isinstance(cantidad_num, float) and math.isnan(cantidad_num)
        if cantidad_num == 0 or es_nan:
            continue

        # ✅ NUEVO: sumar cantidades por categoría (keyword directo)
        if "excav" in desc_norm:
            totales_cant["Excavaciones"] += float(cantidad_num)
        if "rellen" in desc_norm:
            totales_cant["Rellenos"] += float(cantidad_num)
        if re.search(r"\bmr\b", desc_norm):
            totales_cant["Concreto MR"] += float(cantidad_num)
        if "estamp" in desc_norm:
            totales_cant["Concreto estampado"] += float(cantidad_num)

        # ✅ celda real para marcar color
        celda_valor_unit = ws[f"{col_valor}{fila}"]

        # ==========================================
        #  MODO CRÍTICO: SOLO DICCIONARIO (SIN BD)
        # ==========================================
        if modo_critico:
            crit_k, valor_ref = _buscar_critico(desc_norm)
            if valor_ref is None:
                continue  # no es crítica

        # ==========================================
        #  MODO NORMAL: BD (lookup exacto)
        # ==========================================
        else:
            if not unidad:
                continue
            key = (desc_norm, unidad)
            valor_ref = valores_referencia.get(key)
            if valor_ref is None:
                debug["sin_valor_referencia"] += 1
                continue


        # Comparación (rojo si acta > ref, azul si acta < ref)
        TOL = 0.0
        diff = float(valor) - float(valor_ref)

        if abs(diff) > 1:
            if diff > TOL:
                celda_valor_unit.font = Font(color="FFFF0000")  # ROJO
            elif diff < -TOL:
                celda_valor_unit.font = Font(color="FF0000FF")  # AZUL

            valor_ajustado = float(valor_ref) * cantidad_num

            base_registro.append({
                "anio": anio_actual,
                "mes": mes_nombre,
                "archivo": os.path.basename(path_archivo),
                "contratista": nombre_contratista,
                "item": item,
                "descripcion": descripcion,
                "valor_unitario_original": valor,
                "un": unidad,
                "valor_pactado": float(valor_ref),
                "cantidad_presenta": cantidad_num,
                "valor_ajustado": valor_ajustado,
                "modo": "CRITICO" if modo_critico else "NORMAL",
            })
            debug["registros_guardados"] += 1


    # ✅ NUEVO: guardar totales por acta para consolidar luego por contratista
    base_cantidades.append({
        "anio": anio_actual,
        "mes": mes_nombre,
        "archivo": os.path.basename(path_archivo),
        "contratista": nombre_contratista,
        "Excavaciones": totales_cant["Excavaciones"],
        "Rellenos": totales_cant["Rellenos"],
        "Concreto MR": totales_cant["Concreto MR"],
        "Concreto estampado": totales_cant["Concreto estampado"],
        "modo": "CRITICO" if modo_critico else "NORMAL",
    })

    nombre_salida = os.path.join(
        carpeta_salida_mes,
        os.path.basename(path_archivo).replace(".xlsx", "_verificado.xlsx")
    )

    # (lo que ya tenías) hoja de cuadro de cantidades en el verificado
    tablas = _extraer_cantidades_por_familia(ws_vals, columnas)
    _crear_hoja_cuadro_cantidades(wb, tablas, nombre="CUADRO_CANTIDADES")
    
    print("🐞 DEBUG revisar_acta:", os.path.basename(path_archivo), debug)
    
    wb.save(nombre_salida)
    print(f"✔ Revisado: {os.path.basename(path_archivo)}")




